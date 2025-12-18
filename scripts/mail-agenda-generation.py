from datetime import datetime, date
import pandas as pd
import os
import sys
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from exception_list import accountability_exception_list
import warnings
warnings.filterwarnings("ignore")

def formatNumber(number):
    formatted_number = f"{float(number):,.2f}"
    return formatted_number

def checkIfBankIsInGroupedInvoices(bank_code, array):
    for item in array:
        if item['bank_code'] == bank_code:
            return True
    return False

#PESC y UKLY

def transformStringToHTML(bank_code, invoice_num, invoice_month, total_invoice_amount, string, pending_invoices_df, client_name, currency, language):
    #turn format 2409 to September 2024
    year = str(invoice_month)[:4]
    month = str(invoice_month)[4:]
    date_string = ''
    
    month_map = {
        '01': 'January',
        '02': 'February',
        '03': 'March',
        '04': 'April',
        '05': 'May',
        '06': 'June',
        '07': 'July',
        '08': 'August',
        '09': 'September',
        '10': 'October',
        '11': 'November',
        '12': 'December'
    }

    if month in month_map:
        date_string = f'{month_map[month]} {year}'
    else:
        date_string = f'{invoice_month}' #!Caso que nunca deberia pasar



    if language == 'ENG':
        string = f'Dear {client_name},\n\n' + string
    else:
        string = f'Estimado {client_name},\n\n' + string
    new_string = string.replace('\n', '<br>')
    new_string = f'<p class="editor-paragraph">{new_string}</p>'

    if bank_code == 'PESC':
        new_string += f"""
            <p>Por favor asociar esta factura al:</p>
            <ul>
                <li>Contrato: CO-000004776</li>
                <li>Und. Especializada:  OPE-Banca de Servicios</li>
                <li>Grupo: OPE_Distribucion&Activación</li>
            </ul>
            <br>
        """

    if bank_code == 'UKLY':
        new_string += f"""
            <p class="editor-paragraph">
                LLOYDS BANKING GROUP <br>
                Accounts Payable <br>
                BX1 1LT <br>
                Great Britain
            </p>

            <p class="editor-paragraph">
                Vendor name: Intelica Corp <br>
                Vendor no.: 170426 <br>
                PO no.: P019212 <br>
                Invoice no.: {invoice_num} - {date_string} maintenance fee<br>
                Amount: {currency} {total_invoice_amount}
            </p>
            <br>
        """


    if language == 'ENG':
        new_string += "Pending payment invoices: <br>"
    else:
        new_string += "Facturas pendientes de pago: <br>"

    if len(pending_invoices_df) == 0:
        if language == 'ENG':
            new_string += "All submitted invoices have been paid"
        else:
            new_string += "Todas las facturas enviadas han sido pagadas"
    else:
        new_string += f"""
                <br><table border="1" cellpadding="10" cellspacing="0" style="border-collapse:collapse; table-layout:auto; text-align:center;">
                    <tr style="background-color:#555; color:white;">
            """

        if language == 'ENG':
            new_string += f"""
                    <th style="border-right: 1px solid white;">Invoice</th>
                    <th style="border-left: 1px solid white; border-right: 1px solid white;">Date (DD/MM/YYYY)</th>
                    <th style="border-left: 1px solid white; border-right: 1px solid white; padding-left: 15px; padding-right: 15px;">Service Description</th>
                    <th style="border-left: 1px solid white; border-right: 1px solid white; padding-left: 20px; padding-right: 20px;">Amount {currency}</th>
                    <th style="border-left: 1px solid white;">Overdue Days</th>
                </tr>
            """
        else:
            new_string += f"""
                    <th style="border-right: 1px solid white;">Factura</th>
                    <th style="border-left: 1px solid white; border-right: 1px solid white;">Fecha (DD/MM/YYYY)</th>
                    <th style="border-left: 1px solid white; border-right: 1px solid white; padding-left: 15px; padding-right: 15px;">Descripción del Servicio</th>
                    <th style="border-left: 1px solid white; border-right: 1px solid white; padding-left: 20px; padding-right: 20px;">Monto {currency}</th>
                    <th style="border-left: 1px solid white;">Dias vencido</th>
                </tr>
            """

        for index, row in pending_invoices_df.iterrows():
            overdueDaysStyle = ''
            if row['Dias vencido'] > 0:
                #overdueDaysStyle = 'background-color: #FEE7EF; color: #C20E4D;'
                overdueDaysStyle = 'color: red; font-weight: bold;'
            else:
                overdueDaysStyle = ''

            if language == 'ENG':
                new_string += f"""
                <tr>
                    <td>{row['No.']}</td>
                    <td>{row['Date']}</td>
                    <td style="padding-left: 15px; padding-right: 45px; text-align: left;">{row['Memo'] if not pd.isnull(row['Memo']) else 'No description'}</td>
                    <td style="padding-left: 25px; padding-right: 15px; text-align: right;">{formatNumber(row['Amount'])}</td>
                    <td style="{overdueDaysStyle}">{row['Dias vencido']}</td>
                </tr>
                """
            else:
                new_string += f"""
                <tr>
                    <td>{row['No.']}</td>
                    <td>{row['Date']}</td>
                    <td style="padding-left: 15px; padding-right: 45px; text-align: left;">{row['Memo'] if not pd.isnull(row['Memo']) else 'Sin descripción'}</td>
                    <td style="padding-left: 25px; padding-right: 15px; text-align: right;">{formatNumber(row['Amount'])}</td>
                    <td style="{overdueDaysStyle}">{row['Dias vencido']}</td>
                </tr>
                """
        new_string += "</table>"

    if language == 'ENG':
        new_string += f'<br><br><p class="editor-paragraph">Regards,<br>Mauricio<br>Finance Coordinator</p>'
    else:
        new_string += f'<br><br><p class="editor-paragraph">Saludos,<br>Mauricio<br>Finance Coordinator</p>'

    return new_string

def printLog(log):
    print(log)
    sys.stdout.flush()

english_text_content = "I've attached the invoice for the services provided.\nPlease let me know if you have any questions or require further information.\nThank you for your time and attention to this matter.\n\n"
spanish_text_content = "Te adjunto la factura por el servicio prestado.\nPor favor me comenta cualquier duda o detalle.\n\n"

client_master_path = sys.argv[1]
invoice_master_path = sys.argv[2]
new_invoices_path = sys.argv[3]
pending_invoices_path = sys.argv[4]
output_folder = sys.argv[5]
date_string = sys.argv[6]


client_master_df = pd.read_excel(client_master_path)
invoice_master_df = pd.read_excel(invoice_master_path)
new_invoices_df = pd.read_csv(new_invoices_path)
pending_invoices_df = pd.read_excel(pending_invoices_path, skiprows=1)
mail_df = pd.DataFrame(columns=['ID', 'Bank Name', 'Invoice Number', 'Currency', 'Total Amount', 'Mail Recipient', 'Mail CC', 'Mail Title', 'Mail Content', 'Mail Invoices', 'Mail Date', 'Has been sent'])


clients = client_master_df['Bank Code'].unique()
client_data = {}
for client in clients:
    current_client_df = client_master_df[client_master_df['Bank Code'] == client]
    current_invoice_df = invoice_master_df[invoice_master_df['Invoice ID'] == current_client_df['ID'].values[0]]
    current_invoice_date = new_invoices_df['*Invoice Date'].values[0]
    
    if pd.isnull(current_client_df['Mail Recipient'].values[0]):
        continue



    inv_curr_month = current_invoice_date
    inv_curr_month = inv_curr_month.split('/')[2] + inv_curr_month.split('/')[1].zfill(2) if inv_curr_month != '' else ''

    inv_emis_month = current_invoice_df['20240701-NewMonth'].values[0] if not pd.isnull(current_invoice_df['20240701-NewMonth'].values[0]) else ''
    inv_emis_month = str(inv_emis_month)

    current_ = int(inv_curr_month)
    base_ = int(f"202601")
    diff = current_ - base_

    if int(inv_emis_month[4:]) + diff > 12:
        inv_emis_month = str(int(inv_emis_month[:4]) + 1) + str(int(inv_emis_month[4:]) + diff - 12).zfill(2)
    else:
        inv_emis_month = str(int(inv_emis_month[:4])) + str(int(inv_emis_month[4:]) + diff).zfill(2)


    
    client_data[client] = {
        'ids': current_client_df['ID'].values.tolist(),     #get all related IDS in order
        'bank_code': client,
        'bank_country': current_client_df['Country 2'].values[0],
        'bank_name': current_client_df['Bank Name'].values[0],
        'language': current_client_df['Language'].values[0],
        'mail_recipient': current_client_df['Mail Recipient'].values[0],
        'mail_cc': current_client_df['Mail CC'].values[0].split('\n') if not pd.isnull(current_client_df['Mail CC'].values[0]) else [],
        'mail_content': english_text_content if current_client_df['Language'].values[0] == 'ENG' else spanish_text_content,
        'currency': current_client_df['Currency'].values[0],
        'client_name': current_client_df['Client Name'].values[0] if not pd.isnull(current_client_df['Client Name'].values[0]) else 'Client',
        'invoice_emission_month': inv_emis_month,
    }


#Agrupamos los invoices por cliente
invoices = new_invoices_df['Invoice No.'].unique()
grouped_invoices = []
for invoice in invoices:
    curr_bank_code = invoice.split('-')[0].strip()
    if checkIfBankIsInGroupedInvoices(curr_bank_code, grouped_invoices) == False:
        curr_client_data = client_data.get(curr_bank_code)
        if curr_client_data == None:
            print(f" - ERROR - Client {curr_bank_code} not found in client master")
            continue

        summed_amount =  new_invoices_df[new_invoices_df['Invoice No.'] == invoice]['*Amount'].sum()

        current_invoice_clientId = []
        if not invoice[-1].isdigit():
            current_invoice_clientId.append(invoice.split('-')[0].strip() + invoice[-1].strip())
        else:
            current_invoice_clientId = curr_client_data['ids']


        grouped_invoices.append({
            'bank_code': curr_bank_code,
            'ids': current_invoice_clientId,
            'invoice_codes': [invoice],
            'invoice_amt_number': summed_amount,
            'total_invoice_amount': formatNumber(summed_amount)
        })
    else:
        for grouped_invoice in grouped_invoices:
            if grouped_invoice['bank_code'] == curr_bank_code:

                current_ids = grouped_invoice['ids']
                if not invoice[-1].isdigit():
                    current_ids.append(invoice.split('-')[0].strip() + invoice[-1].strip())
                else:
                    current_ids = curr_client_data['ids']

                grouped_invoice['ids'] = current_ids
                grouped_invoice['invoice_codes'].append(invoice)
                grouped_invoice['invoice_amt_number'] += new_invoices_df[new_invoices_df['Invoice No.'] == invoice]['*Amount'].sum()
                grouped_invoice['total_invoice_amount'] = formatNumber(grouped_invoice['invoice_amt_number'])
                break



# def createExcelTable(dest_path, df, check_exists):
#     if check_exists == True and os.path.exists(dest_path):
#         print(f" - MESSAGE - {os.path.basename(dest_path)} file already exists. If you want to overwrite it, please delete it and run the script again")
#     else:
#         writer = pd.ExcelWriter(dest_path, engine='xlsxwriter')
#         df.to_excel(writer, sheet_name='Hoja1', startrow=1, header=False, index=False)
        
#         workbook = writer.book
#         worksheet = writer.sheets['Hoja1']
#         (max_row, max_col) = df.shape

#         column_settings = []
#         for header in df.columns:
#             column_settings.append({'header': header})

#         worksheet.add_table(0, 0, max_row, max_col - 1, {'name': 'Tabla1', 'columns': column_settings})
#         worksheet.set_column(0, max_col - 1, 30)
#         writer.close()
#         print(f' - {os.path.basename(dest_path)} file created successfully')
# createExcelTable(os.path.join(output_folder, 'EMAIL_AGENDA.xlsx'), mail_df, True)



def write_df_as_table(ws, df, start_row, start_col, table_name, fit_content=True):
    # Define header and alternating row colors
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")  # Dark blue
    header_font = Font(bold=True, color="FFFFFF")  # Bold white text
    row_fill_gray = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Darker gray
    row_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White

    # Write dataframe headers with styles
    for col_idx, col_name in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write dataframe rows with alternating colors
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=start_row + 1):
        for c_idx, value in enumerate(row, start=start_col):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            # Apply alternating row colors
            cell.fill = row_fill_gray if (r_idx % 2 == 0) else row_fill_white

    # Adjust column widths to fit content
    if fit_content == True:
        for col_idx in range(start_col, start_col + len(df.columns)):
            col_letter = ws.cell(row=start_row, column=col_idx).column_letter  # Get the column letter
            max_length = max(
                len(str(ws.cell(row=r_idx, column=col_idx).value or "")) for r_idx in range(start_row, start_row + len(df) + 1)
            )
            ws.column_dimensions[col_letter].width = max_length + 8  # Add padding for better spacing

    # Define the table range
    end_row = start_row + len(df)
    end_col = start_col + len(df.columns) - 1
    table_ref = f"{ws.cell(row=start_row, column=start_col).coordinate}:{ws.cell(row=end_row, column=end_col).coordinate}"

    # Create a table with a predefined style
    table = Table(displayName=table_name, ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleMedium9", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    table.tableStyleInfo = style

    ws.add_table(table)
    return end_row



#Generate report based on new_invoices_df
expense_by_client = {}
unique_invoices = new_invoices_df['Invoice No.'].unique()
invoice_summary_df = pd.DataFrame(columns=['#', 'Country', 'Sales Person', 'Invoice No.', 'Client Name', 'Currency', 'Total Amount'])
invoice_summary_per_currency_df = pd.DataFrame(columns=['Currency', 'Invoices', 'Total Amount'])
for idx, invoice in enumerate(unique_invoices):
    invoice_data = new_invoices_df[new_invoices_df['Invoice No.'] == invoice]
    curr_client_data = client_master_df[client_master_df['Bank Code'] == invoice.split('-')[0].strip()]

    invoice_id = invoice.split('-')[0].strip() + invoice.split('-')[1][-1].strip() if not invoice.split('-')[1][-1].isdigit() else invoice.split('-')[0].strip()
    total_amount = 0

    if invoice_id in [item['ID'] for item in accountability_exception_list]:
        for item in accountability_exception_list:
            if invoice_id == item['ID']:
                for idx2, row in invoice_data.iterrows():
                    if float(row['*Amount']) == float(item['AccountabilityAmount']):
                        total_amount = float(item['ServiceAmount'])
                    else:
                        total_amount += float(row['*Amount'])
    else:
        total_amount = invoice_data['*Amount'].sum()

    invoice_summary_df = invoice_summary_df._append({
        '#': idx + 1,
        'Country': curr_client_data['Country 2'].values[0],
        'Sales Person': curr_client_data['Sales Person'].values[0],
        'Invoice No.': invoice,
        'Client Name': curr_client_data['Bank Name'].values[0],
        'Currency': curr_client_data['Currency'].values[0],
        'Total Amount': formatNumber(total_amount)
    }, ignore_index=True)

    #we add to expense_by_client
    bank_code = invoice.split('-')[0].strip()
    if bank_code not in expense_by_client:
        expense_by_client[bank_code] = {
            'amount': total_amount,
            'currency': curr_client_data['Currency'].values[0]
        }
    else:
        expense_by_client[bank_code]['amount'] += total_amount


    #if currency is not present in involice_summary_per_currency_df add it as new row, else add +1 to its invoices and sum the total amount
    if curr_client_data['Currency'].values[0] not in invoice_summary_per_currency_df['Currency'].values:
        invoice_summary_per_currency_df = invoice_summary_per_currency_df._append({
            'Currency': curr_client_data['Currency'].values[0],
            'Invoices': 1,
            'Total Amount': total_amount
        }, ignore_index=True)
    else:
        index = invoice_summary_per_currency_df[invoice_summary_per_currency_df['Currency'] == curr_client_data['Currency'].values[0]].index[0]
        invoice_summary_per_currency_df.at[index, 'Invoices'] += 1
        invoice_summary_per_currency_df.at[index, 'Total Amount'] += total_amount
invoice_summary_per_currency_df['Total Amount'] = invoice_summary_per_currency_df['Total Amount'].apply(lambda x: formatNumber(x))


# Create a workbook and add a worksheet
wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

end_row_1 = write_df_as_table(ws, invoice_summary_df, start_row=1, start_col=1, table_name="Table1")
write_df_as_table(ws, invoice_summary_per_currency_df, start_row=end_row_1 + 3, start_col=1, table_name="Table2")
wb.save(os.path.join(output_folder, 'INVOICE_SUMMARY.xlsx'))



sum_by_currency = pd.DataFrame(columns=['Currency', 'Total Amount'])
for invoice in grouped_invoices:
    print(' - Processing row for client ' + invoice['bank_code'])

    current_client_data = client_data.get(invoice['bank_code'])
    if current_client_data == None:
        print(f" - ERROR - Client {invoice['bank_code']} not found in client master")
        continue

    invoice_title_prefix = 'Intelica Invoice' if current_client_data['language'] == 'ENG' else 'Factura Intelica'
    invoice_file_names = []
    for idx, diff_invoice in enumerate(invoice['invoice_codes']):
        country = client_master_df[client_master_df['ID'] == invoice['ids'][idx]]['Country'].values[0] if not pd.isnull(client_master_df[client_master_df['ID'] == invoice['ids'][idx]]['Country'].values[0]) else ''
        added_country  = f' - {country}' if country != '' else ''
        invoice_file_names.append(invoice['bank_code'][:2] + " - " + invoice_title_prefix + " - " + current_client_data['bank_name'] + " " + diff_invoice + added_country + ".pdf")


    client_pending_invoices = pending_invoices_df[
        (pending_invoices_df['No.'].str.split('-').str[0] == invoice['bank_code']) & 
        (pending_invoices_df['No.'].isin(invoice['invoice_codes']) == False)
    ]


    invoice_number_for_title = invoice['invoice_codes'][0]
    if not invoice_number_for_title[-1].isdigit():
        invoice_number_for_title = invoice_number_for_title[:-1]


    mail_df = mail_df._append({
        'ID': current_client_data['bank_code'],
        'Country': current_client_data['bank_country'],
        'InvoicesAmount': len(current_client_data['ids']),
        'Bank Name': current_client_data['bank_name'],
        'Invoice Number': invoice_number_for_title,
        'Currency': current_client_data['currency'],
        'Total Amount': f'{formatNumber(expense_by_client[invoice["bank_code"]]["amount"])}',
        'Mail Recipient': current_client_data['mail_recipient'],
        'Mail CC': ','.join(current_client_data['mail_cc']),
        'Mail Title': f'Intelica - monthly service invoice {invoice_number_for_title}' if current_client_data['language'] == 'ENG' else f'Intelica - factura mensual de servicios {invoice_number_for_title}',
        'Mail Content': transformStringToHTML(current_client_data['bank_code'], diff_invoice, current_client_data['invoice_emission_month'], invoice['total_invoice_amount'], current_client_data['mail_content'], client_pending_invoices, current_client_data['client_name'], current_client_data['currency'], current_client_data['language']),
        'Mail Invoices': ','.join(invoice_file_names),
        'Mail Date': date_string,
        'Has been sent': "No"
    }, ignore_index=True)

    if current_client_data['currency'] not in sum_by_currency['Currency'].values:
        sum_by_currency = sum_by_currency._append({
            'Currency': current_client_data['currency'],
            'Total Amount': expense_by_client[invoice["bank_code"]]["amount"]
        }, ignore_index=True)
    else:
        index = sum_by_currency[sum_by_currency['Currency'] == current_client_data['currency']].index[0]
        sum_by_currency.at[index, 'Total Amount'] += expense_by_client[invoice["bank_code"]]["amount"]
sum_by_currency['Total Amount'] = sum_by_currency['Total Amount'].apply(lambda x: formatNumber(x))

wb = Workbook()
ws = wb.active
ws.title = "Hoja1"
ws2 = wb.create_sheet("Sheet2")
ws2.title = "Hoja2"

write_df_as_table(ws, mail_df, start_row=1, start_col=1, table_name="Tabla1", fit_content=False)

#add column 'Emails Sent' with value 0 to all rows for invoice_summary_per_currency_df
invoice_summary_per_currency_df['Emails Sent'] = 0

write_df_as_table(ws2, invoice_summary_per_currency_df, start_row=1, start_col=1, table_name="Tabla2", fit_content=True)

if not os.path.exists(os.path.join(output_folder, 'EMAIL_AGENDA.xlsx')):
    wb.save(os.path.join(output_folder, 'EMAIL_AGENDA.xlsx'))
else:
    print(f" - MESSAGE - EMAIL_AGENDA.xlsx file already exists. If you want to overwrite it, please delete it and run the script again")